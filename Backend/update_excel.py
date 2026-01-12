#------------------------------------------------------------------------- Step 1. Improt Libraries ------------------------------------------------------
import sys                                                                  # System library - used to manipulate Python's module search paths (sys.path)   
import os                                                                   # Operating system interface - used for file path operations and directory navigation
import xlwings as xw                                                       # Excel automation library - allows Python to control Excel through COM interface
import pandas as pd                                                       # Data analysis library - used for data manipulation with DataFrames


#------------------------------------------------------ Step 1.1 Display settings (for debugging / console output) ------------------------------------------------------
pd.set_option('display.max_columns', None)                                # When printing DataFrames to console, show ALL columns (don't truncate)
pd.set_option('display.width', None)                                      # Disable line wrapping in console output (show wide tables without breaks)


#--------------------------------------------------------------------- Step 1.2 Path setup --------------------------------------------------------------------------------
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..")) # Get the absolute path of the directory containing this script
scripts_dir = os.path.join(project_root, "Backend")                           # Create path to Backend folder by joining project root with "Backend"

# Add project_root to Python's module search path if not already there
if project_root not in sys.path:                                              # This allows importing modules from the project root
    sys.path.append(project_root)
if scripts_dir not in sys.path:                                                # This allows importing modules from the Backend folder
    sys.path.append(scripts_dir)


#---------------------------------------------------------------- Step 2 Import custom helper functions ---------------------------------------------------------------------
try:
    # Import specific functions from fetch_data.py module
    from fetch_data import (
        fetch_stock_data_with_indicators,  # Function to get stock data with technical indicators
        get_tickers_from_excel,            # Function to read ticker symbols from Excel
        read_manual_esg                    # Function to read ESG data from Manual_ESG sheet
    )
    print("✅ Successfully imported fetch_data functions")
except ImportError:
    # If import fails, print error and exit the program
    print("❌ Could not import fetch_data.")
    sys.exit()  # Terminate the program with exit code 1


#-------------------------------------------------------------------- Step 3 Excel formatting function ------------------------------------------------------------------------
def format_excel(sheet):
    """
    Applies visual formatting to Excel sheet without changing data values.
    This is presentation-only formatting for better readability.
    """
    print("🎨 Applying Excel formatting...")

    # Import get_column_letter function from openpyxl
    # Converts column number (1, 2, 3) to Excel letters (A, B, C)
    from openpyxl.utils import get_column_letter

    # --------------------------------------------------
    # Setup - Get sheet dimensions and headers
    # --------------------------------------------------

    # Read header names from first row: expand("right") gets all cells in row 1
    # .value gets the cell value for each cell
    # Creates list of column names like ["Ticker", "Current Price", ...]
    headers = [c.value for c in sheet.range("A1").expand("right")]

    # Find last used row: expand("down") from A1 gets all cells in column A
    # .last_cell gets the last cell in that range, .row gets its row number
    last_row = sheet.range("A1").expand("down").last_cell.row

    # Find last used column: expand("right") from A1 gets all cells in row 1
    # .last_cell gets the last cell, .column gets its column number
    last_col = sheet.range("A1").expand("right").last_cell.column
    
    # Check if there's actual data (not just header row)
    if last_row < 2:
        print("⚠️ Not enough data to format")
        return  # Exit function early if no data rows
    
    # Define the entire used range from A1 to last cell
    # (row, column) format: (1, 1) = A1
    used_range = sheet.range((1, 1), (last_row, last_col))

    # Helper function: Get the data range for a specific column by name
    def col_range(col_name):
        try:
            # Find column index (1-based) by looking up column name in headers
            idx = headers.index(col_name) + 1
            # Return range from row 2 (first data row) to last_row in that column
            return sheet.range((2, idx), (last_row, idx))
        except ValueError:
            # Column name not found in headers
            print(f"⚠️ Column '{col_name}' not found in headers")
            return None

    # Helper function: Get Excel column letter (A, B, C) for a column name
    def col_letter(col_name):
        try:
            # headers.index gives 0-based index, +1 for 1-based, convert to letter
            return get_column_letter(headers.index(col_name) + 1)
        except ValueError:
            # Column name not found in headers
            print(f"⚠️ Column '{col_name}' not found in headers")
            return None

    # --------------------------------------------------
    # BASE FORMATTING – Apply black theme to all cells
    # --------------------------------------------------

    # Set black background for entire used range
    # 0x000000 is hex for black color
    used_range.api.Interior.Color = 0x000000
    
    # Set white text color for all cells
    # 0xFFFFFF is hex for white color
    used_range.api.Font.Color = 0xFFFFFF
    
    # Set font size to 10 points
    used_range.api.Font.Size = 10
    
    # Center align text horizontally
    # -4108 is Excel constant for xlCenter
    used_range.api.HorizontalAlignment = -4108
    
    # Center align text vertically
    used_range.api.VerticalAlignment = -4108

    # Add grid borders to all cells
    # border_id 7-12 correspond to different border positions (top, bottom, left, right, etc.)
    for border_id in range(7, 13):
        # LineStyle = 1 means continuous line
        used_range.api.Borders(border_id).LineStyle = 1
        
        # Weight = 1 means thin border
        used_range.api.Borders(border_id).Weight = 1
        
        # Set border color to dark gray (0x404040) for subtle contrast on black
        used_range.api.Borders(border_id).Color = 0x404040

    # Auto-fit column widths to content
    used_range.columns.autofit()

    # --------------------------------------------------
    # HEADER ROW STYLING
    # --------------------------------------------------

    # Get header range (all cells in first row)
    header = sheet.range("A1").expand("right")
    
    # Make header text bold
    header.api.Font.Bold = True
    
    # Set header font size slightly larger than data
    header.api.Font.Size = 11
    
    # Set header text color to white
    header.api.Font.Color = 0xFFFFFF
    
    # Set header background color to dark blue (0x2E75B5)
    header.api.Interior.Color = 0x2E75B5

    # Freeze panes for better navigation
    # Freeze header row (row 1) and first 2 columns
    try:
        # Get the active window of Excel
        window = sheet.api.Application.ActiveWindow
        
        # Split after row 1 (freeze row 1)
        window.SplitRow = 1
        
        # Split after column 2 (freeze columns A and B)
        window.SplitColumn = 2
        
        # Actually freeze the panes
        window.FreezePanes = True
    except:
        # If freezing fails, continue without it (non-critical)
        pass

    # --------------------------------------------------
    # TICKER COLUMN SPECIAL STYLING
    # --------------------------------------------------

    # Get column letter for "Ticker" column
    ticker_col = col_letter("Ticker")
    if ticker_col:
        # Get the actual range object for Ticker column (rows 2 to last_row)
        ticker_range = sheet.range(
            (2, headers.index("Ticker") + 1),  # Start at row 2, Ticker column
            (last_row, headers.index("Ticker") + 1)  # End at last row, Ticker column
        )
        # Make ticker symbols bold
        ticker_range.api.Font.Bold = True
        # Set ticker text color to bright blue (0x4FC3F7) for visibility on black
        ticker_range.api.Font.Color = 0x4FC3F7

    # --------------------------------------------------
    # PRICE MOVEMENT INDICATORS FORMATTING
    # --------------------------------------------------

    # Dictionary defining which columns to format and how
    # Format: "Column Name": (good_operator, good_formula, bad_operator, bad_formula)
    price_indicators = {
        "Dividend Yield": (5, ">0", 1, "<0"),  # Operator 5 = greater than, 1 = less than
        "Upside %": (5, ">0", 1, "<0"),        # Same logic for upside percentage
    }
    
    # Loop through each price indicator column
    for col, (good_op, good_formula, bad_op, bad_formula) in price_indicators.items():
        # Get the data range for this column
        rng = col_range(col)
        if rng:  # Only proceed if column exists
            try:
                # Clear any existing conditional formatting from this column
                rng.api.FormatConditions.Delete()

                # Add conditional formatting for POSITIVE values
                # Type=2 means format based on cell value
                # Operator=good_op (5 = greater than)
                # Formula1=">0" means apply formatting when value > 0
                cond = rng.api.FormatConditions.Add(
                    Type=2, Operator=good_op, Formula1=good_formula
                )
                # Set font color to bright green for positive values
                cond.Font.Color = 0x4CAF50
                # Make positive values bold
                cond.Font.Bold = True

                # Add conditional formatting for NEGATIVE values
                # Operator=bad_op (1 = less than)
                # Formula1="<0" means apply formatting when value < 0
                cond = rng.api.FormatConditions.Add(
                    Type=2, Operator=bad_op, Formula1=bad_formula
                )
                # Set font color to bright red for negative values
                cond.Font.Color = 0xF44336
                # Make negative values bold
                cond.Font.Bold = True
            except Exception as e:
                # If formatting fails, print error but continue with other columns
                print(f"⚠️ Error formatting {col}: {e}")

    # --------------------------------------------------
    # TECHNICAL INDICATORS FORMATTING - RSI
    # --------------------------------------------------

    # Get range for RSI (14) column
    rsi_rng = col_range("RSI (14)")
    if rsi_rng:  # Only proceed if RSI column exists
        try:
            # Clear existing conditional formatting
            rsi_rng.api.FormatConditions.Delete()

            # Add formatting for OVERSOLD condition (RSI < 30)
            # Operator=1 means less than
            # Formula1="30" means apply when value < 30
            cond1 = rsi_rng.api.FormatConditions.Add(Type=2, Operator=1, Formula1="30")
            # Set font color to bright blue for oversold
            cond1.Font.Color = 0x2196F3

            # Add formatting for OVERBOUGHT condition (RSI > 70)
            # Operator=5 means greater than
            # Formula1="70" means apply when value > 70
            cond2 = rsi_rng.api.FormatConditions.Add(Type=2, Operator=5, Formula1="70")
            # Set font color to orange for overbought
            cond2.Font.Color = 0xFF9800
            
            # Note: RSI between 30-70 remains white (default color)
            
        except Exception as e:
            # If formatting fails, print error but continue
            print(f"⚠️ Error formatting RSI: {e}")

    # Note: The actual code would continue with SMA, MACD, Analyst Sentiment,
    # Target Prices, Growth Metrics, ESG Scores, Confidence Level, 
    # Categorical columns, Date columns, and Text columns formatting
    # Each follows the same pattern:
    # 1. Get column range with col_range()
    # 2. Clear existing formatting with .FormatConditions.Delete()
    # 3. Add new conditional formatting rules
    # 4. Set font colors for different conditions
    
    print("✅ Black background Excel formatting applied")


#---------------------------------------------------------------- Step 4 Presentation logic for ESG stacking ------------------------------------------------------------------------
def collapse_duplicate_ticker_rows(sheet):
    """
    When a ticker appears multiple times (for different ESG themes):
    - Keep only ESG-related columns on repeated rows
    - Blank all other non-ESG columns for visual clarity
    - This creates a stacked/grouped appearance in the spreadsheet
    """
    
    # Get all column headers from first row
    headers = [cell.value for cell in sheet.range("A1").expand("right")]
    
    # Find last row with data
    last_row = sheet.range("A1").expand("down").last_cell.row

    # Find column index (1-based) of "Ticker" column
    ticker_col_idx = headers.index("Ticker") + 1

    # Define which columns to KEEP when collapsing duplicate rows
    # These are ESG-specific columns that should remain visible
    ESG_ONLY_COLUMNS = {
        "ESG Theme", "Manual ESG Score", "Confidence Level",
        "Assessment Criteria", "Review Date", "Analyst Notes",
        "Upside Bucket", "ESG Category", "RSI Status",
    }

    # Variable to track previous row's ticker
    prev_ticker = None

    # Loop through all data rows (starting at row 2, skipping header)
    for row in range(2, last_row + 1):
        # Get ticker value from current row
        current_ticker = sheet.cells(row, ticker_col_idx).value

        # Check if this ticker is same as previous row's ticker
        if current_ticker == prev_ticker:
            # This is a duplicate ticker row (same stock, different ESG theme)
            
            # Loop through all columns in this row
            for col_idx, col_name in enumerate(headers, start=1):
                # If column is NOT ESG-related AND NOT the Ticker column
                if col_name not in ESG_ONLY_COLUMNS and col_name != "Ticker":
                    # Clear the cell value (set to empty string)
                    sheet.cells(row, col_idx).value = ""

            # Also clear the ticker cell in duplicate rows
            # This creates visual grouping where only first row shows ticker
            sheet.cells(row, ticker_col_idx).value = ""
        else:
            # This is a new ticker, update prev_ticker for next iteration
            prev_ticker = current_ticker


#---------------------------------------------------------------- Step 5 Calculated columns - DataFrame operations ------------------------------------------------------------------------
def add_upside_bucket(df):
    """
    Categorizes 'Upside %' values into human-readable buckets.
    Converts percentage strings like '6.16%' into categories like 'Medium (0–10%)'
    """
    # Define classification function
    def classify(val):
        try:
            # Handle missing or N/A values
            if val == "N/A" or pd.isna(val):
                return "N/A"
            
            # Remove % sign and convert to decimal (6.16% → 0.0616)
            num = float(val.replace("%", "")) / 100
            
            # Categorize based on value
            if num >= 0.10:
                return "High (>10%)"
            elif num >= 0:
                return "Medium (0–10%)"
            else:
                return "Negative"
        except Exception:
            # If conversion fails, return N/A
            return "N/A"

    # Apply classification function to each value in "Upside %" column
    df["Upside Bucket"] = df["Upside %"].apply(classify)
    
    # Return modified DataFrame
    return df

def add_esg_category(df):
    """
    Categorizes 'Manual ESG Score' into Good/Average/Poor buckets.
    Converts numeric scores like 75 into categories like 'Good (≥60)'
    """
    def classify(val):
        try:
            # Handle missing or N/A values
            if val == "N/A" or pd.isna(val):
                return "N/A"
            
            # Convert to float
            score = float(val)
            
            # Categorize based on score
            if score >= 60:
                return "Good (≥60)"
            elif score >= 40:
                return "Average (40–59)"
            else:
                return "Poor (<40)"
        except Exception:
            # If conversion fails, return N/A
            return "N/A"

    # Apply classification function
    df["ESG Category"] = df["Manual ESG Score"].apply(classify)
    
    # Return modified DataFrame
    return df

def add_rsi_status(df):
    """
    Categorizes 'RSI (14)' values into Overbought/Oversold/Neutral.
    Converts numeric RSI like 31.11 into categories like 'Neutral'
    """
    def classify(val):
        try:
            # Handle missing or N/A values
            if val == "N/A" or pd.isna(val):
                return "N/A"
            
            # Convert to float
            rsi = float(val)
            
            # Categorize based on RSI value
            if rsi > 70:
                return "Overbought (>70)"
            elif rsi < 30:
                return "Oversold (<30)"
            else:
                return "Neutral"
        except Exception:
            # If conversion fails, return N/A
            return "N/A"

    # Apply classification function
    df["RSI Status"] = df["RSI (14)"].apply(classify)
    
    # Return modified DataFrame
    return df


#-------------------------------------------------------------------- Step 6 MAIN EXECUTION FUNCTION ------------------------------------------------------------------------
def update_excel():
    """
    Main orchestration function that runs the entire pipeline:
    1. Locate and open Excel file
    2. Fetch data from various sources
    3. Merge and enrich data
    4. Write to Excel
    5. Apply formatting
    6. Collapse duplicate rows for presentation
    """
    try:
        print("🔄 Starting Excel update...")

        # --------------------------------------------------
        # STEP 1: Locate Excel file
        # --------------------------------------------------
        
        # Get base directory (one level above this script)
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        
        # Construct full path to Excel file
        excel_path = os.path.join(base_dir, "Stock_data.xlsm")

        # Check if file exists
        if not os.path.exists(excel_path):
            print("❌ Stock_data.xlsm not found")
            return False  # Return failure

        # --------------------------------------------------
        # STEP 2: Connect to Excel application
        # --------------------------------------------------
        
        # Get active Excel application instance
        app = xw.apps.active
        if app is None:
            # No Excel instance is running
            print("❌ Please open Excel first")
            return False

        # --------------------------------------------------
        # STEP 3: Open or attach to workbook
        # --------------------------------------------------
        
        # Look for already open workbook with "Stock_data.xlsm" in name
        # next() returns first matching workbook or None if not found
        workbook = next(
            (wb for wb in app.books if "Stock_data.xlsm" in wb.name),
            None
        )

        # If workbook not already open, open it
        if workbook is None:
            workbook = app.books.open(excel_path)

        # Get the "RawData" sheet
        sheet = workbook.sheets["RawData"]

        # --------------------------------------------------
        # STEP 4: Fetch input data
        # --------------------------------------------------
        
        print("📥 Fetching tickers...")
        # Read ticker symbols from Sheet1 of the Excel file
        tickers = get_tickers_from_excel(excel_path, sheet_name="Sheet1")

        print("📊 Fetching automated data...")
        # Fetch stock data with technical indicators for all tickers
        df = fetch_stock_data_with_indicators(tickers)

        print("📝 Reading Manual_ESG...")
        # Read ESG data from Manual_ESG sheet
        manual_esg_df = read_manual_esg(excel_path)

        # --------------------------------------------------
        # STEP 5: Merge and enrich data
        # --------------------------------------------------
        
        # Merge stock data with ESG data on Ticker column
        # how="left" keeps all stock data rows, adds ESG where available
        final_df = df.merge(
            manual_esg_df,
            on="Ticker",
            how="left"
        )

        # Add calculated columns (buckets/categories)
        final_df = add_upside_bucket(final_df)
        final_df = add_esg_category(final_df)
        final_df = add_rsi_status(final_df)

        # --------------------------------------------------
        # STEP 6: Write to Excel
        # --------------------------------------------------
        
        print("✍️ Writing to Excel...")
        
        # Clear existing content in RawData sheet
        sheet.clear()
        
        # Write DataFrame to sheet starting at cell A1
        # xlwings automatically writes headers and all data
        sheet.range("A1").value = final_df

        # --------------------------------------------------
        # STEP 7: Apply formatting
        # --------------------------------------------------
        
        # Apply visual formatting (colors, borders, alignment)
        format_excel(sheet)

        # Apply presentation logic (collapse duplicate ticker rows)
        collapse_duplicate_ticker_rows(sheet)
        
        # --------------------------------------------------
        # STEP 8: Completion
        # --------------------------------------------------
        
        print("✅ Excel update complete")
        return True  # Return success

    except Exception as e:
        # If ANY error occurs in the try block
        #print(f"❌ Error: {e}")
        
        # Print detailed traceback for debugging
        # import traceback
        # traceback.print_exc()
        
        return False  # Return failure

#-------------------------------------------------------------------- Step 7 SCRIPT ENTRY POINT ------------------------------------------------------------------------
if __name__ == "__main__":                                                          # Run only if file is executed directly
    update_excel()                                                                   # Call the main function